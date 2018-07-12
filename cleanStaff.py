from openpyxl import Workbook
import json


def clean_staff():



    # Inicializar listado en jsonxx
    with open('dealerCodes.json', 'r') as f:
        dealer_dict = json.load(f)


    # Inicializar excel workook
    wb = Workbook()
    ws = wb.active
    record = ""
    count = 0
    row_line = 1
    stafffile = input("Escriba el nombre del archivo: ")
    # Encargado de leer el archivo de staffmaster
    staff_namefile = './input/' + stafffile + '.txt'
    staff_textfile = open(staff_namefile, 'r')

    # Excel Header
    column_a = ws.cell(row=1, column=1)
    column_a.value = "Dealer Code"
    column_b = ws.cell(row=1, column=2)
    column_b.value = "Spin"
    column_c = ws.cell(row=1, column=3)
    column_c.value = "Name"
    column_d = ws.cell(row=1, column=4)
    column_d.value = "Middle Name"
    column_e = ws.cell(row=1, column=5)
    column_e.value = "Last Name"
    column_f = ws.cell(row=1, column=6)
    column_f.value = "Job Code"
    column_g = ws.cell(row=1, column=7)
    column_g.value = "Department"
    column_h = ws.cell(row=1, column=8)
    column_h.value = "Hired Date"

    for line in staff_textfile:
        line = line.split("|")
        has_term_date = (line[8][0:1]).isdigit()
        has_t = line[9].startswith("T")
        have_blank_fields = (line[5] + " " + line[6] + " " + line[7] + " " + line[8] + " " + line[9]).strip()
        txt_line = line[0]

        for emp_code_json in dealer_dict['dealer_codes']:
            emp_code_json = str(emp_code_json['code'])

            if txt_line == emp_code_json:
                record = line[0]

        if not has_t and not has_term_date and have_blank_fields != "" and record != line[0]:

            #print(line)
            count = count + 1

            row_line = row_line + 1
            #send data to Excel
            a2 = ws.cell(row=row_line, column=1)
            a2.value = line[0]
            b2 = ws.cell(row=row_line, column=2)
            b2.value = line[1]
            c2 = ws.cell(row=row_line, column=3)
            c2.value = line[2]
            d2 = ws.cell(row=row_line, column=4)
            d2.value = line[3]
            e2 = ws.cell(row=row_line, column=5)
            e2.value = line[4]
            f2 = ws.cell(row=row_line, column=6)
            f2.value = line[5]
            g2 = ws.cell(row=row_line, column=7)
            g2.value = line[6]
            h2 = ws.cell(row=row_line, column=8)
            h2.value = line[7]

    print("Cantidad de records: " + str(count))
    wb.save(stafffile + ".xlsx")

# STAFFMASTER - 20180629


clean_staff()
